"""
Import historical PM data from data/PASS_Inventory.xlsx into the inventory DB.

Re-running this command is safe (idempotent) — uses get_or_create / update_or_create
for every row, so it can be run again after fixing the source file.
"""

from collections import defaultdict
from pathlib import Path

import openpyxl
from django.conf import settings
from django.db import transaction

from django.core.management.base import BaseCommand

from inventory.models import (
    AssetType, Device, Plant, PlantRoundNote, PMRound, PMSnapshot, SnPrefix,
)

EXCEL_PATH = Path(settings.BASE_DIR) / "data" / "PASS_Inventory.xlsx"

PLANTS = [
    ("PKT", "ภูเก็ต", "Phuket"),
    ("HDY", "หาดใหญ่", "Hat Yai"),
    ("KBV", "กระบี่", "Krabi"),
    ("URT", "สุราษฎร์ธานี", "Surat Thani"),
    ("UTP", "อู่ตะเภา", "U-Tapao"),
    ("UDT", "อุดรธานี", "Udon Thani"),
    ("CNX", "เชียงใหม่", "Chiang Mai"),
]

# prefix -> ปีรุ่น (จาก legend แถว 2 ในไฟล์ Excel) — REC ไม่อยู่ใน legend, เว้นไว้
SN_PREFIXES = [
    ("RP9", 2025),
    ("RNC", 2023),
    ("RPB", 2023),
    ("RMA", 2021),
    ("RJB", 2018),
]

# (year, quarter, field_start, field_end, document_due)
# วันที่ปี 2025 ตามแผนจริง — 2024/2026 copy เดือน/วันเดิม (ปรับได้ทีหลังใน admin)
PM_ROUNDS = [
    (2024, "Q1", "2024-02-03", "2024-03-22", "2024-03-31"),
    (2024, "Q2", "2024-05-31", "2024-07-19", "2024-07-31"),
    (2024, "Q3", "2024-09-13", "2024-11-09", "2024-11-21"),
    (2025, "Q1", "2025-02-03", "2025-03-22", "2025-03-31"),
    (2025, "Q2", "2025-05-31", "2025-07-19", "2025-07-31"),
    (2025, "Q3", "2025-09-13", "2025-11-09", "2025-11-21"),
    (2026, "Q1", "2026-02-03", "2026-03-22", "2026-03-31"),
    (2026, "Q2", "2026-05-31", "2026-07-19", "2026-07-31"),
    (2026, "Q3", "2026-09-13", "2026-11-09", "2026-11-21"),
]
ROUND_ORDER = [(year, quarter) for year, quarter, *_ in PM_ROUNDS]

# Excel column letters: (S/N column, หมายเหตุ column) ต่อรอบ
ROUND_COLUMNS = {
    (2024, "Q1"): ("B", "C"),
    (2024, "Q2"): ("D", "E"),
    (2024, "Q3"): ("F", "G"),
    (2025, "Q1"): ("H", "I"),
    (2025, "Q2"): ("J", "K"),
    (2025, "Q3"): ("L", "M"),
    (2026, "Q1"): ("O", "P"),
    (2026, "Q2"): ("Q", "R"),
    (2026, "Q3"): ("S", "T"),
}
HW_COLUMN = "N"  # หมายเหตุระดับ plant+รอบ (2025 Q3) — merged cell คลุมหลายแถว
HW_ROUND = (2025, "Q3")
DATA_ROWS = range(6, 48)  # แถวข้อมูลอุปกรณ์ (6-47)

# S/N เดิมพิมพ์ผิด ถูกแก้ในไฟล์ด้วยหมายเหตุ "แก้ไขจากเลข ..." — ถือเป็นเครื่องเดียวกัน
SN_RENAME = {
    "RMA03F1034": "RMA03F1035",
}

# เครื่องใหม่ -> เครื่องเก่าที่ถูกแทน (retired กับ new_delivery ในแถว/รอบติดกัน)
REPLACEMENTS = {
    "RP903T0403": "RMA03F1028",
    "RP903T0400": "RJB03F0083",
    "RPB03T0403": "RJB03F0088",
    "RPB03T0404": "REC39F1523",
    "RP903T0393": "RJB03F0078",
    "RP903T0399": "RJB03F0089",
    "RPB03T0407": "REC39F1528",
    "RPB03T0408": "REC39F1533",
    "RPB03T0406": "RJB03F0084",
    "RP903T0394": "REC39F1527",
}


def classify(note):
    """แปลงข้อความหมายเหตุดิบ -> PMSnapshot.Status (ตาม mapping ใน 02_schema_design.md)"""
    note = note or ""
    if "เสื่อมสภาพ" in note:
        return PMSnapshot.Status.RETIRED
    if "ส่งมอบ" in note:
        return PMSnapshot.Status.NEW_DELIVERY
    if "ยืมจาก" in note:
        return PMSnapshot.Status.BORROWED_IN
    if "ยืมใช้" in note:
        return PMSnapshot.Status.LOANED_OUT
    if "แก้ไขจากเลข" in note:
        return PMSnapshot.Status.CORRECTION
    return PMSnapshot.Status.IN_USE


class Command(BaseCommand):
    help = "Import historical PM snapshots from data/PASS_Inventory.xlsx"

    @transaction.atomic
    def handle(self, *args, **options):
        plants = self._load_plants()
        asset_type = self._load_asset_type()
        rounds = self._load_pm_rounds()
        self._load_sn_prefixes()

        raw_entries, hw_notes = self._read_sheet(plants)
        devices_data = self._resolve_devices(raw_entries)
        n_devices, n_snapshots = self._save(devices_data, asset_type, rounds, hw_notes, plants)

        self.stdout.write(self.style.SUCCESS(
            f"Imported {n_devices} devices, {n_snapshots} PM snapshots, "
            f"{len(hw_notes)} plant round note(s)."
        ))

    # ── Master data ─────────────────────────────────────────────────────────
    def _load_plants(self):
        plants = {}
        for code, name_th, name_en in PLANTS:
            plant, _ = Plant.objects.update_or_create(
                code=code, defaults={"name_th": name_th, "name_en": name_en}
            )
            plants[code] = plant
        return plants

    def _load_asset_type(self):
        asset_type, _ = AssetType.objects.update_or_create(
            code="TABLET", defaults={"name": "Tablet", "default_warranty_months": 12}
        )
        return asset_type

    def _load_pm_rounds(self):
        rounds = {}
        for year, quarter, start, end, due in PM_ROUNDS:
            pm_round, _ = PMRound.objects.update_or_create(
                year=year, quarter=quarter,
                defaults={
                    "field_start_date": start,
                    "field_end_date": end,
                    "document_due_date": due,
                },
            )
            rounds[(year, quarter)] = pm_round
        return rounds

    def _load_sn_prefixes(self):
        for prefix, year in SN_PREFIXES:
            SnPrefix.objects.update_or_create(prefix=prefix, defaults={"year": year})

    # ── Read Excel grid ─────────────────────────────────────────────────────
    def _read_sheet(self, plants):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb["Sheet1"]

        raw_entries = defaultdict(list)  # sn -> [(plant_code, round_key, note), ...]
        hw_notes = {}  # plant_code -> note (2025 Q3 H/W)
        current_plant = None

        for row in DATA_ROWS:
            plant_name = ws[f"A{row}"].value
            if plant_name:
                current_plant = self._plant_code_by_name(plant_name, plants)

            hw_value = ws[f"{HW_COLUMN}{row}"].value
            if hw_value and hw_value.strip():
                hw_notes[current_plant] = hw_value.strip()

            for round_key, (sn_col, note_col) in ROUND_COLUMNS.items():
                sn = ws[f"{sn_col}{row}"].value
                if not sn:
                    continue
                sn = str(sn).strip()
                sn = SN_RENAME.get(sn, sn)
                note = ws[f"{note_col}{row}"].value
                note = str(note).strip() if note else ""
                raw_entries[sn].append((current_plant, round_key, note))

        return raw_entries, hw_notes

    def _plant_code_by_name(self, name_th, plants):
        for code, plant_name_th, _ in PLANTS:
            if plant_name_th == name_th:
                return code
        raise ValueError(f"Unknown plant name in Excel: {name_th!r}")

    # ── Resolve borrow situations -> (home_plant, snapshots) per device ────
    def _resolve_devices(self, raw_entries):
        devices = {}
        for sn, entries in raw_entries.items():
            by_plant = defaultdict(dict)
            for plant_code, round_key, note in entries:
                by_plant[plant_code][round_key] = (note, classify(note))

            if len(by_plant) == 1:
                home_plant, rounds_data = next(iter(by_plant.items()))
                final = {
                    rk: (home_plant, note, status)
                    for rk, (note, status) in rounds_data.items()
                }
            else:
                # ยืมข้าม plant: ฝั่งที่มี status loaned_out = เจ้าของ (home_plant)
                # ฝั่งที่เหลือ = ผู้ยืม -> ใช้เป็น snapshot จริงตาม D1
                (plant_a, data_a), (plant_b, data_b) = by_plant.items()
                if any(status == PMSnapshot.Status.LOANED_OUT for _, status in data_a.values()):
                    owner, owner_data, borrower, borrower_data = plant_a, data_a, plant_b, data_b
                else:
                    owner, owner_data, borrower, borrower_data = plant_b, data_b, plant_a, data_a

                home_plant = owner
                final = {}
                for rk in set(owner_data) | set(borrower_data):
                    if rk in borrower_data:
                        note, status = borrower_data[rk]
                        final[rk] = (borrower, note, status)
                    else:
                        note, status = owner_data[rk]
                        final[rk] = (owner, note, status)

            devices[sn] = {"home_plant": home_plant, "snapshots": final}
        return devices

    # ── Persist ──────────────────────────────────────────────────────────────
    def _save(self, devices_data, asset_type, rounds, hw_notes, plants):
        device_objs = {}
        n_snapshots = 0

        for sn, data in devices_data.items():
            snapshots_sorted = sorted(
                data["snapshots"].items(), key=lambda kv: ROUND_ORDER.index(kv[0])
            )
            received_round = next(
                (rounds[rk] for rk, (_, _, status) in snapshots_sorted
                 if status == PMSnapshot.Status.NEW_DELIVERY),
                None,
            )
            last_round_key, (_, _, last_status) = snapshots_sorted[-1]
            is_retired = last_status == PMSnapshot.Status.RETIRED
            retired_round = rounds[last_round_key] if is_retired else None

            device, _ = Device.objects.update_or_create(
                serial_number=sn,
                defaults=dict(
                    asset_type=asset_type,
                    home_plant=plants[data["home_plant"]],
                    received_round=received_round,
                    is_retired=is_retired,
                    retired_round=retired_round,
                ),
            )
            device_objs[sn] = device

            for round_key, (plant_code, note, status) in snapshots_sorted:
                PMSnapshot.objects.update_or_create(
                    device=device,
                    pm_round=rounds[round_key],
                    defaults=dict(plant=plants[plant_code], status=status, note=note),
                )
                n_snapshots += 1

        # Replacement links — ทำหลังสร้าง Device ครบทุกตัว
        for new_sn, old_sn in REPLACEMENTS.items():
            if new_sn in device_objs and old_sn in device_objs:
                device_objs[new_sn].replaced_predecessor = device_objs[old_sn]
                device_objs[new_sn].save(update_fields=["replaced_predecessor"])

        # Plant round notes (H/W)
        for plant_code, note in hw_notes.items():
            PlantRoundNote.objects.update_or_create(
                plant=plants[plant_code], pm_round=rounds[HW_ROUND],
                defaults={"note": note},
            )

        return len(device_objs), n_snapshots
